from openpyxl import load_workbook

# import os
# os.environ.setdefault("DJANGO_SETTINGS_MODULE", "oasis2018.settings")
# import django
# django.setup()

from shop.models.wallet import Wallet
from shop.models.balance import Balance
from registrations.models import Bitsian

class Verifier():

    def __init__(self):
        self.valid_set = set()

    def populate_valid_set(self):
        wb = load_workbook("/home/hemanth/Desktop/DVM Oasis/oasis-2018/scripts/student_list3.xlsx")
        ws = wb.active
        for i in range(2, 4664, 1):
            print(ws["J%d" % i].value)
            self.valid_set.add(ws["J%d" % i].value)

    def validate_all_bitsians(self):
        valid_count = 0
        invalid_count = 0
        total_count = 0
        for bitsian in Bitsian.objects.all():
            if bitsian.email in self.valid_set:
                valid_count += 1
            else:
                invalid_count += 1
            total_count += 1
        print("No. of valid bitsians: %d" % valid_count)
        print("No. of invalid bitsians: %d" % invalid_count)
        print("Total no. of bitsians: %d" % total_count)


if __name__ == "__main__":
    v = Verifier()
    v.populate_valid_set
    v.validate_all_bitsians()
