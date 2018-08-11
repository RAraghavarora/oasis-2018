from openpyxl import Workbook
#from events.models import *
from preregistration.models import *
from django.contrib.auth.models import User


def getAll():
    getGenParticipant()
    getRoctaves()
    getRapWarsExtension()
    getPurpleProseExtension()
    getStandupSoapboxExtension()
    getCategory()
    getIntroEvent()
    getParticipation()
    print("GENERATION COMPLETE!")

def getGenParticipant():
    all_gps = GenParticipant.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "genparticipant"

    ws["A1"] = "Name"
    ws["B1"] = "City"
    ws["C1"] = "Phone"
    ws["D1"] = "Gender"
    ws["E1"] = "Email"

    row = 2
    for gp in all_gps:
        ws["A{}".format(row)] = gp.name
        ws["B{}".format(row)] = gp.city
        ws["C{}".format(row)] = gp.phone
        ws["D{}".format(row)] = gp.gender
        ws["E{}".format(row)] = gp.email_address
        row += 1

    wb.save("./scripts/DataSheets/GenParticipant.xlsx")
    print("GenParticipant.xlsx generated.")


def getRoctaves():
    all_rs = Roctaves.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "roctaves"

    ws["A1"] = "Name"
    ws["B1"] = "Genre"
    ws["C1"] = "Email"
    ws["D1"] = "Phone"
    ws["E1"] = "NOP"
    ws["F1"] = "EP"
    ws["G1"] = "e1"
    ws["H1"] = "e2"
    ws["I1"] = "e"

    row = 2
    for rs in all_rs:
        ws["A{}".format(row)] = rs.name
        ws["B{}".format(row)] = rs.genre
        ws["C{}".format(row)] = rs.email_address
        ws["D{}".format(row)] = rs.phone
        ws["E{}".format(row)] = rs.number_of_participants
        ws["F{}".format(row)] = rs.elimination_preference
        ws["G{}".format(row)] = rs.entry1
        ws["H{}".format(row)] = rs.entry2
        ws["I{}".format(row)] = rs.enteries
        row += 1

    wb.save("./scripts/DataSheets/Roctaves.xlsx")
    print("Roctaves.xlsx generated.")


def getRapWarsExtension():
    all_rwe = RapWarsExtension.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "RWE"

    ws["A1"] = "Participant(FK)"
    ws["B1"] = "RN"
    ws["D1"] = "COP"

    row = 2
    for rwe in all_rwe:
        ws["A{}".format(row)] = "{}_+_{}_+_{}_+_{}".format(rwe.participant.name, rwe.participant.phone, rwe.participant.email_address, rwe.participant.city)
        ws["B{}".format(row)] = rwe.rapper_name
        ws["D{}".format(row)] = rwe.city_of_participation
        row += 1

    wb.save("./scripts/DataSheets/RapWarsExtension.xlsx")
    print("RapWarsExtension.xlsx generated.")


def getPurpleProseExtension():
    all_ppe = PurpleProseExtension.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "PPE"

    ws["A1"] = "Participant(FK)"
    ws["B1"] = "College"
    ws["C1"] = "YASOS"
    ws["D1"] = "COP"
    ws["E1"] = "Entry"

    row = 2
    for ppe in all_ppe:
        ws["A{}".format(row)] = "{}_+_{}_+_{}_+_{}".format(ppe.participant.name, ppe.participant.phone, ppe.participant.email_address, ppe.participant.city)
        ws["B{}".format(row)] = ppe.college
        ws["C{}".format(row)] = ppe.year_and_stream_of_study
        ws["D{}".format(row)] = ppe.city_of_participation
        ws["E{}".format(row)] = ppe.entry

        row += 1

    wb.save("./scripts/DataSheets/PurpleProseExtension.xlsx")
    print("PurpleProseExtension.xlsx generated.")


def getStandupSoapboxExtension():
    all_sse = StandupSoapboxExtension.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "SSE"

    ws["A1"] = "Participant(FK)"
    ws["B1"] = "TDS"
    ws["C1"] = "PC"
    ws["D1"] = "COP"

    row = 2
    for sse in all_sse:
        ws["A{}".format(row)] = "{}_+_{}_+_{}_+_{}".format(sse.participant.name, sse.participant.phone, sse.participant.email_address, sse.participant.city)
        ws["B{}".format(row)] = sse.time_doing_standup
        ws["C{}".format(row)] = sse.previous_competition
        ws["D{}".format(row)] = sse.city_of_participation
        row += 1

    wb.save("./scripts/DataSheets/StandupSoapboxExtension.xlsx")
    print("StandupSoapboxExtension.xlsx generated.")

def getCategory():
    all_cat = Category.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Category"

    ws["A1"] = Name

    row=2
    for cat in all_cat:
        ws["A{}".format(row)] = cat.name
        row += 1

    wb.save("./scripts/DataSheets/Category.xlsx")
    print("Category.xlsx generated.")

def getCategory():
    all_cat = Category.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Category"

    ws["A1"] = "Name"

    row=2
    for cat in all_cat:
        ws["A{}".format(row)] = cat.name
        row += 1

    wb.save("./scripts/DataSheets/Category.xlsx")
    print("Category.xlsx generated.")

def getIntroEvent():
    all_ie = IntroEvent.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "IE"

    ws["A1"] = "User(OOF)"
    ws["B1"] = "Name"
    ws["C1"] = "SD"
    ws["D1"] = "Rules"
    ws["E1"] = "Category(FK)"
    ws["F1"] = "Contact"

    row=2
    for ie in all_ie:
        ws["A{}".format(row)] = ie.user.username
        ws["B{}".format(row)] = ie.name
        ws["C{}".format(row)] = ie.short_description
        ws["D{}".format(row)] = ie.rules
        ws["E{}".format(row)] = ie.category.name
        ws["F{}".format(row)] = ie.contact
        row += 1

    wb.save("./scripts/DataSheets/IntroEvent.xlsx")
    print("IntroEvent.xlsx generated.")

def getParticipation():
    all_part = Participation.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Participation"

    ws["A1"] = "Event(FK)"
    ws["B1"] = "Participant(FK)"
    ws["C1"] = "PCR Approved"
    ws["D1"] = "CR Approved"

    row=2
    for part in all_part:
        ws["A{}".format(row)] = part.introevent.name
        ws["A{}".format(row)] = "{}_+_{}_+_{}_+_{}".format(part.participant.name, part.participant.phone, part.participant.email_address, part.participant.city)
        ws["C{}".format(row)] = part.pcr_approved
        ws["D{}".format(row)] = part.cr_approved
        row += 1

    wb.save("./scripts/DataSheets/Participation.xlsx")
    print("Participation.xlsx generated.")
