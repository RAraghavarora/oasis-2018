import sys
from openpyxl import *
filename2='/home/sanchit/DVM_SHIT/oasis-2018/scripts/N2O Signings Final.xlsx'
#f1 = '/home/sanchit/DVM_SHIT/oasis-2018/scripts/FINAL STUDENT LIST.xlsx'
wb = load_workbook(filename=filename2)
sheet = wb["MASTERSHEET"] #Signings
#wb2 = load_workbook(filename=f1)
#sheet2 = wb2['MASTERSHEET'] #Student list
length = len(tuple(sheet.rows))


row=3
l2 = len(tuple(sheet.rows))
for i in range(3,4656):
    s_id=(str(sheet.cell(row=i,column=1).value))
    s_id=s_id.lower()
    try:
        if s_id[1:5]=='2018' or s_id[1:5]=='2017':
            email=s_id[:-1]+'@pilani.bits-pilani.ac.in'
        else:
            email=s_id[0:5]+s_id[6:9]+'@pilani.bits-pilani.ac.in'
    
        sheet["G{}".format(row)] = email    
        row+=1
    except Exception as e:
        print(e,i)


wb.save(filename2)



# emailist=[]
# long_list=[]
# l=[]
# l2 = len(tuple(sheet2.rows))
# p = []
# for j in range(2, l2):
#     idd = sheet2.cell(row=j, column=2).value
#     ee = sheet2.cell(row=j, column=8).value
#     p.append([ee, idd])
# dick = {idd: ee for ee, idd in p}
# #print(len(dick))
# fin = []
# lit = 0
# for i in range(2, length+1):

#     un= str(sheet.cell(row=i,column=1).value)
#     long_id = sheet.cell(row=i, column=1).value
#     try:
#         l.append(dick[long_id])
#         email = dick[long_id]
#         # print(1)
#         fin.append([email])
#         lit+=2
#     except Exception as e:
#         print (3)
#         print(i, e)
#         pass


# import csv
# with open('eggs2.csv', 'w') as f:
#     w = csv.writer(f)
#     for i in fin:
#         w.writerow(i)
# print(len(long_list))




