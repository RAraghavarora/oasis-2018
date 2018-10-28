from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from openpyxl import load_workbook
import csv

filename = BASE_DIR + "/scripts/Guthrie Govan signings.xlsx"
filename2='/home/sanchit/Downloads/Guthrie Govan signings.xlsx'
wb = load_workbook(filename=filename)
sheet = wb["actual"]

length = len(tuple(sheet.rows))

for i in range(2, length+1):

	un= str(sheet.cell(row=i,column=1).value)
	# email = sheet.cell(row=i, column=2).value
	long_id = sheet.cell(row=i, column=2).value
	try:
		if long_id[3] == '7' or long_id[3] == '8':
			email = str(un[0]).lower()
			email = email + (long_id[0:4] + long_id[8:] + '@pilani.bits-pilani.ac.in')
		else:
			email = str(un[0]).lower()
			email = email + (long_id[0:4] + long_id[9:] + '@pilani.bits-pilani.ac.in')

		print(email)
	except Exception as e:
		continue
