import sys
from openpyxl import *
filename2='scripts/IndoSoul.xlsx'
f1 = 'scripts/student_list3.xlsx'
wb = load_workbook(filename=filename2)
sheet = wb["Sheet1"]
wb2 = load_workbook(filename=f1)
sheet2 = wb2['MASTERSHEET'] #Student list
length = len(tuple(sheet.rows))
emailist=[]
l=[]
l2 = len(tuple(sheet2.rows))
p = []
for j in range(2, l2+1):
    bitsid = sheet2.cell(row=j, column=2).value
    email = sheet2.cell(row=j, column=11).value
    p.append([email, bitsid])
dick = {bitsid: email for email, bitsid in p}
#print(len(dick))
a=0
for i in range(3, length+1):

	long_id = sheet.cell(row=i, column=1).value
	try:
		l.append(dick[long_id])
	except Exception as e:
		a+=1

wb.close()
wb2.close()
print(len(l))
for i in l:
	print(i)