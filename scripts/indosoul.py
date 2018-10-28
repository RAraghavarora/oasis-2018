from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from shop.models.item import *
from events.models import *
# from shop.models.transaction import *
from openpyxl import *
# from random import *
filename = BASE_DIR + "/scripts/IndoSoul DVM.xlsx"
wb = load_workbook(filename=filename,data_only=True)
sheet1=wb['Sheet1']
length1 = len(tuple(sheet1.rows))

for i in range(3,length1):
    bits_id1=sheet1.cell(row=i,column=1).value
    # count=int(sheet1.cell(row=i,column=3).value)
    # print(count)
    # if count==1:
    #     bitsian=Bitsian.objects.get(long_id=bits_id1)
    #     Tickets.objects.get_or_create(prof_show=prof_show,user=bitsian.user,count=1,is_excel_sheet=True)
    #     print("obj "+str(i))
        
    # elif count==2:
       
    #    bitsian=Bitsian.objects.get(long_id=bits_id1)
    #    Tickets.objects.get_or_create(prof_show=prof_show,user=bitsian.user,count=2,is_excel_sheet=True)
    #    print("obj "+str(i))
        
    # else:
    #     continue








# profshow=MainProfShow.objects.get(id=7)