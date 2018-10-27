from registrations.models import Bitsian
from django.contrib.auth.models import User
from openpyxl import *
from random import *
_dir='/home/sanchit/Downloads/'
wb = load_workbook(_dir + 'MESSLIST.xlsx')
sheet = wb['Sheet1']

def fun():
    for i in range(100,1000):
        bits_id=sheet.cell(row=i,column=1).value    
        name=sheet.cell(row=i,column=2).value
        sex=sheet.cell(row=i,column=3).value
        bhawan=sheet.cell(row=i,column=4).value
        #room=int(sheet.cell(row=i,column=5).value)    
        username='f'+(bits_id[0:4]+bits_id[8:12])
        chars = 'abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
        password = ''.join(choice(chars) for _ in range(8))
        user=User.objects.create(username=username,password='')
        user.set_password(password)
        user.save()
        email=username+'@pilani.bits-pilani.ac.in'
        barcode=username#change this to make a proper one
        Bitsian.objects.get_or_create(barcode=barcode,name=name,long_id=bits_id,gender=sex,bhawan=bhawan,user=user,email=email
        )
        print("obj"+str(i))

fun()