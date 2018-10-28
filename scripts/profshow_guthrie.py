from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from shop.models.item import *
from events.models import *
# from shop.models.transaction import *
from openpyxl import *
# from random import *
filename = BASE_DIR + "/scripts/Guthrie Govan signings.xlsx"
wb = load_workbook(filename=filename,data_only=True)
sheet2=wb['Google form']
sheet1 = wb['actual']



length1 = len(tuple(sheet1.rows))
length2 = len(tuple(sheet2.rows))
lista=[]
listb=[]
for i in range(2,length1+1):
    bits_id1=sheet1.cell(row=i,column=2).value
    lista.append(bits_id1)

for i in range(2,length2+1):
    bits_id2=sheet2.cell(row=i,column=4).value
    bits_id2=bits_id2[:-1]
    listb.append(bits_id2)
 
listfinal=list(set(lista+listb))

cnt=0
for bits_id in listfinal:
    try:
        
        bitsian=Bitsian.objects.get(long_id=bits_id)
        profshow=MainProfShow.objects.get(name='Guthrie Govan')
        try:
            t = Tickets.objects.get(prof_show=profshow,user=bitsian.user)
            t.count+=1
            t.save()
        except:
            t=Tickets.objects.create(prof_show=profshow,user=bitsian.user, count=1)

        cnt+=1
        print(bitsian.user.username,t.count, str(cnt))
    except (Bitsian.DoesNotExist, MainProfShow.DoesNotExist):
        pass
    except Exception as e:
        print(e)





    
    
   
    


