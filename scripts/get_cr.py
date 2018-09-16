from openpyxl import Workbook
from registrations.models import IntroReg

def getCollegeRepresentative():
    objects = IntroReg.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "College Representative"

    ws["A1"] = "Name"
    ws["B1"] = "College"
    ws["C1"] = "Mobile"
    ws["D1"] = "Email"

    row = 2
    for gp in objects:
        ws["A{}".format(row)] = gp.name
        ws["B{}".format(row)] = gp.college.name
        ws["C{}".format(row)] = gp.mobile_no
        ws["D{}".format(row)] = gp.email_id
        row += 1

    wb.save("./scripts/DataSheets/CollegeRepresentative.xlsx")
    print("CollegeRepresentative.xlsx generated.")
