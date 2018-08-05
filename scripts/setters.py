from openpyxl import load_workbook
from events.models import *
from preregistration.models import *
from django.contrib.auth.models import User


def setAll():
    setGenParticipant()
    setRoctaves()
    setRapWarsExtension()
    setPurpleProseExtension()
    setStandupSoapboxExtension()
    setCategory()
    setIntroEvent()
    setParticipation()
    print("ALL SET!")

def setGenParticipant():
    wb = load_workbook('./scripts/DataSheets/GenParticipant.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            name = row[0].value
            city = row[1].value
            phone = row[2].value
            gender = row[3].value
            email = row[4].value
            GenParticipant.objects.get_or_create(name=name, city=city, phone=phone, gender=gender, email_address=email)
        else:
            initial = False
    print("GenParticipant set.")

def setRoctaves():
    wb = load_workbook('./scripts/DataSheets/Roctaves.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            name = row[0].value
            genre = row[1].value
            email = row[2].value
            phone = row[3].value
            nop = int(row[4].value)
            ep = row[5].value
            e1 = row[6].value
            e2 = row[7].value
            e = row[8].value
            new = Roctaves.objects.get_or_create(name=name, genre=genre, email_address=email, phone=phone, number_of_participants=nop, elimination_preference=ep, entry1=e1, entry2=e2, enteries=e)
        else:
            initial = False
    print("Roctaves set.")

def setRapWarsExtension():
    wb = load_workbook('./scripts/DataSheets/RapWarsExtension.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            gpname, gpphone, gpemail, gpcity = row[0].value.split("_+_")
            gp = GenParticipant.objects.get(name=gpname, phone=gpphone, email_address=gpemail, city=gpcity)
            rn = row[1].value
            cop = row[3].value
            try:
                RapWarsExtension.objects.get_or_create(participant=gp, rapper_name=rn, city_of_participation=cop)
            except:
                pass
        else:
            initial = False
    print("RapWars set.")

def setPurpleProseExtension():
    wb = load_workbook('./scripts/DataSheets/PurpleProseExtension.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            gpname, gpphone, gpemail, gpcity = row[0].value.split("_+_")
            try:
                gp = GenParticipant.objects.get(name=gpname, phone=gpphone, email_address=gpemail, city=gpcity)
            except:
                gp = GenParticipant.objects.filter(name=gpname, phone=gpphone, email_address=gpemail)[0]
            college = row[1].value
            yasos = row[2].value
            cop = row[3].value
            entry = row[4].value
            try:
                PurpleProseExtension.objects.get_or_create(participant=gp, college=college, year_and_stream_of_study=yasos, city_of_participation=cop, entry=entry)
            except:
                pass
        else:
            initial = False
    print("PurpleProseExtension set.")

def setStandupSoapboxExtension():
    wb = load_workbook('./scripts/DataSheets/StandupSoapboxExtension.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            gpname, gpphone, gpemail, gpcity = row[0].value.split("_+_")
            try:
                gp = GenParticipant.objects.get(name=gpname, phone=gpphone, email_address=gpemail, city=gpcity)
            except:
                gp = GenParticipant.objects.filter(name=gpname, phone=gpphone, email_address=gpemail)[0]
            tds = row[1].value
            pc = row[2].value
            cop = row[3].value
            try:
                StandupSoapboxExtension.objects.get_or_create(participant=gp, time_doing_standup=tds, previous_competition=pc, city_of_participation=cop)
            except:
                pass
        else:
            initial = False
    print("StandupSoapboxExtension set.")

def setCategory():
    wb = load_workbook('./scripts/DataSheets/Category.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            Category.objects.get_or_create(name=row[0].value)
        else:
            initial = False
    print("Category set.")

def setUsers():
    try:
        User.objects.get(username='Eminem')
    except:
        User.objects.create_user(username='Eminem', password='Rap1ikeyoumean!t')
    try:
        User.objects.get(username='Rocky')
    except:
        User.objects.create_user(username='Rocky', password='B@$kin101')
    try:
        User.objects.get(username='Shaky')
    except:
        User.objects.create_user(username='Shaky', password='BThou$hallnotpasskin101')

def setIntroEvent():
    setUsers()
    wb = load_workbook('./scripts/DataSheets/IntroEvent.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            try:
                user = User.objects.get(username=row[0].value)
            except:
                user = User.objects.create_user(username=row[0].value, password="asdfghjkl")
            name = row[1].value
            sd = row[2].value
            if sd == None:
                sd = ""
            rules = row[3].value
            if rules == None:
                rules = ""
            cat = Category.objects.get(name=row[4].value)
            contact = row[5].value
            IntroEvent.objects.get_or_create(user=user, name=name, short_description=sd, rules=rules, category=cat, contact=contact)
        else:
            initial = False
    print("IntroEvent set.")

def setParticipation():
    wb = load_workbook('./scripts/DataSheets/Participation.xlsx')
    ws = wb.active
    initial = True # a switch for skipping the first row
    for row in ws.rows:
        if not initial:
            ie = Introevent.objects.get(name=row[0].value)
            gpname, gpphone, gpemail, gpcity = row[1].value.split("_+_")
            try:
                gp = GenParticipant.objects.get(name=gpname, phone=gpphone, email_address=gpemail, city=gpcity)
            except:
                gp = GenParticipant.objects.filter(name=gpname, phone=gpphone, email_address=gpemail)[0]
            participations.objects.create(event=introevent, participant=gp, pcr_approved=row[2].value, cr_approved=row[3].value)
        else:
            initial = False
    print("participation set.")
