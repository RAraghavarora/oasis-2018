from openpyxl import *
from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from shop.models.item import *
from events.models import *
from shop.models.transaction import *
from random import *
filename = BASE_DIR + "/scripts/N2O Signings Final.xlsx"
#filename2='/home/sanchit/DVM_SHIT/oasis-2018/scripts/N2O Signings Final.xlsx'
wb = load_workbook(filename=filename,data_only=True)
sheet1=wb['MASTERSHEET']
# length1 = len(tuple(sheet1.rows))

cnt=0
for i in range(3,905):
    bitsmail_id1=sheet1.cell(row=i,column=7).value #mailid1
    if sheet1.cell(row=i,column=6).value:
        count=int(sheet1.cell(row=i,column=6).value)
         
        try:
            bitsian=Bitsian.objects.get(email=bitsmail_id1)
            profshow=MainProfShow.objects.get(name="N20")
            try:
                t = Tickets.objects.get(prof_show=profshow,user=bitsian.user)
                t.count+=count
                t.save()
            except:
                Tickets.objects.create(prof_show=profshow,user=bitsian.user, count=count)
            cnt+=1
            print("obj "+str(cnt))
        
        except Exception as e:
            print(e,i)


# profshow=MainProfShow.objects.get(id=7)