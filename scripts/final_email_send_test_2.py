from django.contrib.auth.models import User
from registrations.models import *
from events.models import *
import sys
from shop.models.item import Tickets
from openpyxl import *
from oasis2018.settings import BASE_DIR

filename=BASE_DIR + '/scripts/IndoSoul.xlsx'
print(1)
wb = load_workbook(filename=filename)
print(2)
sheet = wb["Sheet1"]
print(3)
length = 4662
print(4)
profshow = MainProfShow.objects.get(name__icontains='Indosoul')
print("RAGHAV")
# index_count = 0
sum = 0
for i in range(3,length+1):
	try:
		name = sheet.cell(row=i,column=2).value
		long_id = sheet.cell(row=i,column=1).value
		email= sheet.cell(row=i, column=9).value
		ticket_count = sheet.cell(row=i,column=8).value
		# print(type(ticket_count))
		if not ticket_count:
			continue
			# ticket_count=0
		try:
			ticket_count = int(ticket_count)
		except:
			ticket_count = None
			continue
			print(type(ticket_count), ticket_count, "!")

		sum+=ticket_count
		username = str(email).split('@')[0]

		user,created = User.objects.get_or_create(username=username)
		# print(user)

		if created:
			user.set_password('abcxyz123' + str(i))
			user.save()

		try:
			bitsian,created=Bitsian.objects.get_or_create(user=user,long_id=long_id,name=name)
		except Exception as e:
			print(e)
			print(user)
			continue
		if created:
			bitsian.save()

		try:
			ticket = Tickets.objects.get(prof_show=profshow,user=user)
			ticket.count+=ticket_count
			ticket.save()
			print(i)
		except:
			print(i)
			ticket = Tickets.objects.create(prof_show=profshow, user=user, count = ticket_count)
		# index_count+=1
		# print('Object '+str(i))	

	except Exception as error:
		print(str(i)+'\t'+str(error))
		break
print(sum)