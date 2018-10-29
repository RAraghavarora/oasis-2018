import sys
from openpyxl import *
filename2='/home/sanchit/DVM_SHIT/oasis-2018/scripts/SEL and EDM signings sheet Pre-Oasis.xlsx'
f1 = '/home/sanchit/DVM_SHIT/oasis-2018/scripts/FINAL STUDENT LIST.xlsx'
wb = load_workbook(filename=filename2)
sheet = wb["Sheet1"]
wb2 = load_workbook(filename=f1)
sheet2 = wb2['MASTERSHEET'] #Student list
length = len(tuple(sheet.rows))
emailist=[]
long_list=[]
l=[]
l2 = len(tuple(sheet2.rows))
p = []
for j in range(2, l2):
    idd = sheet2.cell(row=j, column=2).value
    ee = sheet2.cell(row=j, column=8).value
    p.append([ee, idd])
dick = {idd: ee for ee, idd in p}
#print(len(dick))
fin = []
lit = 0
for i in range(2, length+1):

    un= str(sheet.cell(row=i,column=1).value)
    long_id = sheet.cell(row=i, column=1).value
    try:
        l.append(dick[long_id])
        email = dick[long_id]
        # print(1)
        fin.append([email, sheet.cell(row=i, column=8).value, sheet.cell(row=i, column=9).value, sheet.cell(row=i, column=10).value])
        lit+=2
    except Exception as e:
        print (3)
        print(i, e)
        pass
import csv
with open('eggs.csv', 'w') as f:
    w = csv.writer(f)
    for i in fin:
        w.writerow(i)
print(len(long_list))




