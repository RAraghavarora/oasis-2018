from django.shortcuts import render, get_object_or_404, get_list_or_404, redirect
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from registrations.models import *
from .models import *
from events.models import *
from django.contrib.auth.models import User
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import login, logout
from django.core.urlresolvers import reverse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from functools import reduce
from registrations.urls import *
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from oasis2018.settings import BASE_DIR
import os
from time import gmtime, strftime
import string
from pcradmin.views import get_cr_name, get_pcr_number
from django.contrib import messages
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required


import sendgrid
from sendgrid.helpers.mail import *
from . import send_grid

import os
import re

from utils.registrations import get_pcr_number

from oasis2018.settings_config.keyconfig import *
from random import sample, choice
from django.contrib import messages
chars = string.ascii_lowercase + string.ascii_uppercase + string.digits

###Helper function to get the group leader##
def get_group_leader(group):
    try:
        return group.participant_set.get(is_g_leader=True)
    except:
        return None

#To create a code for the group
def generate_group_code(group):
    group_id = group.id
    encoded = group.group_code
    if encoded == '':
        raise ValueError
    if encoded is not None:
        return encoded
    # group_ida = "%04d" % int(group_id)
    group_leader = get_group_leader(group)
    college_code = ''.join(group_leader.college.name[0:3])
    # if len(college_code)<4:
    #     college_code += str(0)*(4-len(college_code))
    college_code = college_code + str(group.id)
    group.group_code = college_code
    # group.group_code = college_code + group_ida
    group.save()
    return encoded

@staff_member_required
def index(request):
	if request.user.username.lower() == 'controlz':
		return redirect(reverse('regsoft:controlz_home'))
	if request.user.username.lower() == 'recnacc':
		return redirect(reverse('regsoft:recnacc_home'))
	if request.user.username == 'firewallz' or request.user.is_superuser:
		return redirect(reverse('regsoft:firewallz_home'))

############FIREWALLZ#############

@staff_member_required
def firewallz_home(request):
    college_list = [college for college in College.objects.all() if college.participant_set.filter(is_cr=True)]

    rows = []
    for college in college_list:
        name = college.name
        # print(college)
        cr = college.participant_set.get(college=college, is_cr=True).name
        total_final = college.participant_set.filter(pcr_final=True).count()
        firewallz_passed = college.participant_set.filter(pcr_final=True, firewallz_passed=True).count()
        url = request.build_absolute_uri(reverse('regsoft:firewallz_approval', kwargs={'c_id':college.id}))
        rows.append({'data': [name,cr,total_final,firewallz_passed] , 'link':[{'url':url,'title':'Approve Participants'},]})

    headings = ['College', 'CR', 'PCr Finalised Participants', 'Firewallz Passed','Approve Participants']
    title = 'Select college to approve Participants'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def firewallz_approval(request, c_id):
    college = get_object_or_404(College, id=c_id)
    if request.method == 'POST':
        try:
            data = request.POST
            id_list = data.getlist('id_list')
        except:
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            g_leader_id = data['g_leader_id']
            if not g_leader_id in id_list:
                messages.warning(request,'Please select a Group Leader from the participant list.')
                return redirect(request.META.get('HTTP_REFERER'))   
            g_leader = Participant.objects.get(id=g_leader_id)
            g_leader.is_g_leader = True
            g_leader.save()
        
        except:
            messages.warning(request,'Please select a Group Leader.')
            return redirect(request.META.get('HTTP_REFERER'))
        
        group = Group.objects.create()
        for part_id in id_list:
            part = Participant.objects.get(id=part_id)
            if part.group is not None:
                g_leader.is_g_leader = False
                g_leader.save()
                group.delete()
                context = {
                    'error_heading': "Error",
                    'message': "Participant(s) already in a group",
                    'url':request.build_absolute_uri(reverse('regsoft:firewallz_home'))
                    }
                return render(request, 'registrations/message.html', context)
            part.firewallz_passed=True
            part.group = group
            part.save()
        encoded = generate_group_code(group)
        group.save()
        part_list = Participant.objects.filter(id__in=id_list)
        return redirect(reverse('regsoft:get_group_list', kwargs={'g_id':group.id}))
        
    groups_passed = [group for group in Group.objects.all() if get_group_leader(group) and get_group_leader(group).college == college]
    unapproved_list = college.participant_set.filter(pcr_final=True, firewallz_passed=False, is_guest=False)
    return render(request, 'regsoft/firewallz_approval.html', {'groups_passed':groups_passed, 'unapproved_list':unapproved_list, 'college':college})

@staff_member_required
def get_group_list(request, g_id):
    try:
        group = get_object_or_404(Group, id=g_id)
    except:
        context = {
            'error_heading': "Error",
            'message': "Group does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:firewallz_home'))
            }
        return render(request, 'registrations/message.html', context)
    if request.method == 'POST':
        try:
            data = request.POST
            id_list = data.getlist('id_list')
        except:
            return redirect(request.META.get('HTTP_REFERER'))

        participant_list = Participant.objects.filter(id__in=id_list)
        for participant in participant_list:
            if participant.is_g_leader:
                messages.warning(request,'Cannot unconfirm a Group Leader.')
                continue
            participant.group = None
            participant.firewallz_passed = False
            participant.save()
        leader = get_group_leader(group)
        if group.participant_set.count == 0:
            group.delete()
        return redirect(reverse('regsoft:get_group_list', kwargs={'g_id':g_id}))
    participant_list = group.participant_set.all()
    return render(request, 'regsoft/group_list.html', {'participant_list':participant_list, 'group':group})

@staff_member_required
def delete_group(request, g_id):
    try:
        group = get_object_or_404(Group, id=g_id)
    except:
        context = {
        'error_heading': "Error",
        'message': "No such group exists",
        'url':request.build_absolute_uri(reverse('regsoft:firewallz_home'))
        }
        return render(request, 'registrations/message.html', context)
    leader = get_group_leader(group)
    for part in group.participant_set.all():
        part.firewallz_passed = False
        part.is_g_leader = False
        part.save()
    group.delete()
    return redirect(reverse('regsoft:firewallz_approval', kwargs={'c_id':leader.college.id}))

@staff_member_required
def add_guest(request):
    if request.method == 'GET':
        colleges = College.objects.all()
        guests = Participant.objects.filter(is_guest=True)
        return render(request, 'regsoft/add_guest.html', {'colleges':colleges, 'guests':guests})

    elif request.method == 'POST':
        data = request.POST.dict()
        email = data['email']
        if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", email):
            messages.warning(request,'Please enter a valid email address.')
            return redirect(request.META.get('HTTP_REFERER'))
        else:
            try:
                Participant.objects.get(email=data['email'])
                messages.warning(request,'Email already registered.')
                return redirect(request.META.get('HTTP_REFERER'))
            except:
                pass
            participant = Participant()
            if not data['name']:
                messages.warning(request, 'Please enter your name')
                return redirect(request.META.get('HTTP_REFERER'))
            if len(data['phone'])==10:
                participant.phone = int(data['phone'])
            else:
                messages.warning(request, 'Please enter a valid phone number')
                return redirect(request.META.get('HTTP_REFERER'))

            participant.name = ' '.join(str(data['name']).strip().split())
            participant.gender = str(data['gender'])
            participant.city = str(data['city'])
            participant.email = str(data['email'])
            participant.college = College.objects.get(name=str(data['college']))
            
            if not re.match(
                r'^20\d{2}(A[1-578B]([PT]S|A[1-578B]|B[1-5])|[CD]2[TP]S|B[1-5]([PT]S|A[1-578B])|H[DS0-9]\d{2}|PH[X0-9][PF0-9])\d{4}P$',
                str(data['bits_id'])):
                messages.warning(request, 'Please enter a proper bits id')
                return redirect(request.META.get('HTTP_REFERER'))
            participant.bits_id = str(data['bits_id'])

            participant.is_guest = True
            participant.email_verified = True
            participant.save()
            participant.firewallz_passed = True
            
            username = participant.name.split(' ')[0] + str(participant.id)
            password = ''.join(choice(chars) for i in range(8)) #random alphanumeric password of length 8
            user = User.objects.create_user(username=username, password=password)
            participant.user = user
            participant.save() # Barcode will automatically be generated and assigned. Django Signals
            
            #sendgrid email body is written in a separate file called send_grid.py
            email_class = send_grid.add_guest() 
            send_to = participant.email
            name = participant.name
            body = email_class.body%(name, username, password, get_pcr_number())
            to_email = Email(send_to)
            content = Content('text/html', body)

            try:
                mail = Mail(email_class.from_email,email_class.subject,to_email,content)
                response = send_grid.sg.client.mail.send.post(request_body = mail.get())
            except:
                participant.user = None
                participant.save()
                user.delete()
                participant.delete()
                context = {
                    'url':request.build_absolute_uri(reverse('regsoft:firewallz_home')),
                    'error_heading': "Error sending mail",
                    'message': "Sorry! Error in sending email. Please try again.",
                }
                return render(request, 'registrations/message.html', context)

            context = {
                'error_heading': "Emails sent",
                'message': "Login credentials have been mailed to the corresponding new participants.",
                'url':request.build_absolute_uri(reverse('regsoft:firewallz_home'))
            }
            return render(request, 'registrations/message.html', context)
    
@staff_member_required
def remove_guests(request):
    if request.method == 'POST':
        data = request.POST
        try:
            list = data.getlist('guest_list')
        except:
            messages.warning(request, 'No guest selected.')
            return redirect(request.META.get('HTTP_REFERER'))
        participants=Participant.objects.filter(id__in=data.getlist('guest_list'), is_guest=True)
        for participant in participants:
            user = participant.user
            participant.user = None
            user.delete()
        participants.delete()
    return redirect(reverse('regsoft:add_guest'))
    
@staff_member_required
def add_participant(request):
    if request.method == 'GET':
        event_list = MainEvent.objects.all()
        colleges = College.objects.all()
        return render(request, 'regsoft/add_participant.html', {'event_list':event_list,'colleges':colleges})

    if request.method == 'POST':
        data = request.POST
        email = data['email']
        if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", email):
            messages.warning(request,'Please enter a valid email address.')
            return redirect(request.META.get('HTTP_REFERER'))
        else:
            try:
                Participant.objects.get(email=data['email'])
                messages.warning(request,'Email already registered.')
                return redirect(request.META.get('HTTP_REFERER'))
            except:
                pass
            try:
                events = data.getlist('events')
            except:
                pass
            participant = Participant()
            if not data['name']:
                messages.warning(request, 'Please enter your name')
                return redirect(request.META.get('HTTP_REFERER'))
            if len(data['phone'])==10:
                participant.phone = int(data['phone'])
            else:
                messages.warning(request, 'Please enter a valid phone number')
                return redirect(request.META.get('HTTP_REFERER'))
            participant.name = str(data['name'])
            participant.phone = int(data['phone'])
            participant.gender = str(data['gender'])
            participant.city = str(data['city'])
            participant.email = str(data['email'])
            participant.college = College.objects.get(name=str(data['college']))
           
            participant.email_verified = True
            participant.pcr_approved = True
            participant.save()
            participant.pcr_final = True

            username = participant.name.split(' ')[0] + str(participant.id)
            password = ''.join(choice(chars) for _ in range(8)) #random alphanumeric password of length 8
            user = User.objects.create_user(username=username, password=password)
            participant.user = user
            participant.save() # Here, barcode will automatically be generated and stored using django signals.
            college = participant.college
            if not college.participant_set.filter(is_cr=True):
                participant.is_cr = True
                participant.save()
            try:
                events = data.getlist('events')
                for key in data.getlist('events'):
                    event = Event.objects.get(id=int(key))
                    MainParticipation.objects.create(event=event, participant=participant, pcr_approved=True)
            except:
                pass
            participant.save()

            #sendgrid email body is written in a separate file called send_grid.py
            #DRY
            email_class = send_grid.add_guest() 
            send_to = participant.email
            name = participant.name
            body = email_class.body%(name, username, password, get_pcr_number())
            to_email = Email(send_to)
            content = Content('text/html', body)

            try:
                mail = Mail(email_class.from_email,email_class.subject,to_email,content)
                response = send_grid.sg.client.mail.send.post(request_body = mail.get())
            except:
                participant.user = None
                participant.save()
                user.delete()
                participant.delete()
                context = {
                    'url':request.build_absolute_uri(reverse('regsoft:firewallz_home')),
                    'error_heading': "Error sending mail",
                    'message': "Sorry! Error in sending email. Please try again.",
                }
                return render(request, 'registrations/message.html', context)
            context = {
                'error_heading': "Emails sent",
                'message': "Login credentials have been mailed to the corresponding new participants.",
                'url':request.build_absolute_uri(reverse('regsoft:firewallz_home'))
            }
            return render(request, 'registrations/message.html', context)
    

@staff_member_required
def approved_groups(request):
    groups_passed = Group.objects.all()
    colleges = [get_group_leader(group).college for group in Group.objects.all()]
    data = [{'college': college, "group": group} for (college, group) in zip(colleges, groups_passed)]
    return render(request, 'regsoft/approved_groups.html', {'data':data})


# #########Recnacc#########

@staff_member_required
def recnacc_home(request):
    try:    
        rows = [{'data':[group.group_code, get_group_leader(group).name, get_group_leader(group).college.name, get_group_leader(group).phone,group.created_time, group.participant_set.filter(controlz=True).count(), group.participant_set.filter(controlz=True, acco=True, checkout_group=None).count(), group.participant_set.filter(checkout_group__isnull=False).count()], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:allocate_participants', kwargs={'g_id':group.id})), 'title':'Allocate Participants'}]} for group in Group.objects.all().order_by('-created_time')]
    except:
        rows =[]        

    title='Groups that have passed Firewallz'
    headings = ['Group Code', 'Group Leader', 'College', 'Gleader phone', 'Firewallz passed time', 'Total controlz passed','Total alloted', 'Checkout','View Participants']

    table={
        'rows':rows,
        'title':title,
        'headings':headings
    }
    return render(request,'regsoft/tables.html',{'tables':[table,]})


@staff_member_required
def allocate_participants(request,g_id):
    group=get_object_or_404(Group,id=g_id)
    if request.method=='POST':
        from datetime import datetime
        data=request.POST
        try:
            group.amount_deduct=data['amount_retained']
            group.save()
        except:
            pass
        if data['action']=='allocate':
            try:
                parts_id=data.getlist('data')
                room_id=data['room']
                room=Room.objects.get(id=room_id)
            except:
                messages.warning(request,'Incomplete selection')
                return redirect(request.META.get('HTTP_REFERER'))
            if not parts_id:
                messages.warning(request,'Incomplete selection')
                return redirect(request.META.get('HTTP_REFERER'))
            rows=[]
            room.vacancy-=len(parts_id)
            room.save()
            for part_id in parts_id:
                part=Participant.objects.get(id=part_id)
                part.acco=True
                part.room=room
                part.recnacc_time=datetime.now()
                part.save()
        elif data['action']=='deallocate':
            try:
                parts_id = data.getlist('data')
            except:
                return redirect(request.META.get('HTTP_REFERER'))
            for part_id in parts_id:
                part = Participant.objects.get(id=part_id)
                room = part.room
                room.vacancy += 1
                room.save()
                part.acco = False
                part.room = None
                part.save()
        return redirect(reverse('regsoft:recnacc_group_list',kwargs={'c_id':get_group_leader(group).college.id}))
    else:
        room_list=Room.objects.all()
        unalloted_participants=group.participant_set.filter(acco=False,checkout_group=None,controlz=True)
        alloted_participants=group.participant_set.filter(acco=True,checkout_group=None,controlz=True)
        checked_out=group.participant_set.filter(checkout_group__isnull=False)
        return render(request,'regsoft/allot.html',{'unalloted':unalloted_participants,'alloted':alloted_participants,'rooms':room_list,'group':group,'checked_out':checked_out
    })


@staff_member_required
def recnacc_group_list(request,c_id):
    college=get_object_or_404(College,id=c_id)
    try:
        group_list=[group for group in Group.objects.all() if get_group_leader(group).college==college]
    except:
        context = {
            'error_heading': "Error",
            'message': "Group leader does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:recnacc_home'))
            }
        return render(request, 'registrations/message.html', context)
    complete_groups = [group for group in group_list if all(part.acco for part in group.participant_set.filter(controlz=True))]
    incomplete_groups=[group for group in group_list if not group in complete_groups]

    complete_rows = [{'data':[group.created_time, get_group_leader(group).name, group.participant_set.filter(controlz=True).count(), group.participant_set.filter(acco=True,checkout_group=None, controlz=True).count()], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:allocate_participants', kwargs={'g_id':group.id})), 'title':'Manage group'}]} for group in complete_groups]

    incomplete_rows = [{'data':[group.created_time, get_group_leader(group).name, group.participant_set.filter(controlz=True).count(), group.participant_set.filter(acco=True,checkout_group=None, controlz=True).count()], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:allocate_participants', kwargs={'g_id':group.id})), 'title':'Manage group'}]} for group in incomplete_groups]

    complete_table={
        'rows':complete_rows,
        'headings':['Created Time','GroupLeader Name','Total Control Passed','Alloted','Manage'],
        'title':'Completely alloted groups from '+college.name
    }
    incomplete_table={
        'rows':incomplete_rows,
        'headings':['Created Time','GroupLeader Name','Total Control Passed','Alloted','Manage'],
        'title':'Incompletely alloted groups from '+college.name
    }
    return render(request,'regsoft/tables.html',{'tables':[complete_table,incomplete_table],'college':college})


@staff_member_required
def room_details(request):
    room_list = Room.objects.all()
    rows = [{'data':[room.room, room.bhavan.name, room.vacancy, room.capacity,], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:manage_vacancies', kwargs={'r_id':room.id})), 'title':'Manage'},]} for room in room_list]
    headings = ['Room', 'Bhavan', 'Vacancy', 'Capacity', 'Manage Room Details']
    title = 'Manage Room Details'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title,
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def manage_vacancies(request,r_id):
    room=get_object_or_404(Room,id=r_id)
    if request.method=='POST':
        data=request.POST
        try:
            vacancy = int(data["vacancy"])
            capacity = int(data["capacity"])
        except KeyError as missing:
            messages.warning(request,'Please enter both vacancy and capacity.')
            return redirect(reverse('regsoft:room_details'))
        if vacancy > capacity:
            messages.warning(request,'Vacancy cannot be greater than capacity.')
            return redirect(reverse('regsoft:room_details'))
        room.vacancy = vacancy
        room.capacity = capacity
        room.save()      
        try:
            note=data['note']
            re_note=Note()
            re_note.note=note
            re_note.room=room
            re_note.save()
        except:
            messages.warning(request,'Please add a note.')      
        # try:
        #     room.vacancy=data['vacancy']
        #     room.save()
        # except:
        #     pass
        # try:
        #     capacity=room.capacity
        #     vacancy=int(data['capacity'])-capacity
        #     room.vacancy=int(data['vacancy'])+vacancy
        #     room.capacity=data['capacity']
        #     room.save()
        # except:
        #     pass
        return redirect(reverse('regsoft:room_details'))
    else:
        notes=room.note_set.all()
        return render(request,'regsoft/manage_vacancies.html',{'room':room,'notes':notes})

@staff_member_required
def recnacc_bhavans(request):
    rows =[{'data':[bhavan.name, reduce(lambda x,y:x+y.vacancy, bhavan.room_set.all(), 0),], 'link':[{'title':'Details', 'url':request.build_absolute_uri(reverse('regsoft:bhavan_details', kwargs={'b_id':bhavan.id}))},] } for bhavan in Bhavan.objects.all()]
    headings = ['Bhavan', 'Vacancy', 'Room-wise details']
    tables = [{'title':'All Bhavans', 'headings':headings, 'rows':rows}]
    return render(request,'regsoft/tables.html', {'tables':tables})

@staff_member_required
def bhavan_details(request, b_id):
    bhavan = Bhavan.objects.get(id=b_id)
    rows = [{'data':[room.room, room.vacancy, room.capacity], 'link':[{'title':'Details', 'url':request.build_absolute_uri(reverse('regsoft:manage_vacancies', kwargs={'r_id':room.id}))},]} for room in bhavan.room_set.all()]
    headings = ['Room', 'Vacancy', 'Capacity', 'Manage Vacancies']
    tables = [{'title': 'Details for ' + bhavan.name + ' bhavan', 'headings':headings, 'rows':rows}]
    return render(request, 'regsoft/tables.html', {'tables':tables})

@staff_member_required
def group_vs_bhavan(request):
    rows = []
    for group in Group.objects.all():
        if group.participant_set.filter(acco=True):
            rooms = []
            for part in group.participant_set.filter(acco=True):
                
                if not part.room in rooms:
                    room = part.room 
                    rooms.append(room)
            for room in rooms:
                rows.append({'data':[room,get_group_leader(group).college.name,group.group_code, get_group_leader(group).name, group.participant_set.filter(acco=True, room=room).count(), get_group_leader(group).phone],'link':[]})
    table = {
        'rows':rows,
        'headings':['Room','College','Group Code', 'Group Leader', 'Number of participants in bhavan', 'Group Leader Phone'],
        'title':'Group vs Bhavans'
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})


@staff_member_required
def recnacc_college_details(request):
    college_list = []
    for c in College.objects.all():
        try:
            p = c.participant_set.get(is_cr=True)
            college_list.append(c)
        except:
            pass
    rows = [{'data':[college.name, college.participant_set.get(is_cr=True).name,college.participant_set.filter(acco=True).count()], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:recnacc_group_list', kwargs={'c_id':college.id})), 'title':'View Details'}]} for college in college_list]
    headings = ['College', 'Cr Name','Alloted Participants', 'View Details']
    title = 'Select college to approve Participants'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def checkout_college(request):
    rows = [{'data':[college.name,college.participant_set.filter(acco=True).count(), college.participant_set.filter(checkout_group__isnull=False).count()],'link':[{'title':'Checkout', 'url':request.build_absolute_uri(reverse('regsoft:checkout', kwargs={'c_id':college.id}))}] } for college in College.objects.all()]
    tables = [{'title':'List of Colleges', 'rows':rows, 'headings':['College', 'Alloted Participants', 'Checked out Participants','Checkout']}]
    return render(request, 'regsoft/tables.html', {'tables':tables})

###Helper function to generate checkout group code###
def generate_ckgroup_code(group):
    group_id=group.id
    encoded=group.group_code
    if encoded=='':
        raise ValueError
    if encoded is not None:
        return encoded
    group_ida = "%04d" % int(group_id)
    college_code = ''.join(group.participant_set.all()[0].college.name.split(' '))
    if len(college_code)<4:
        college_code += str(0)*(4-len(college_code))
    group.group_code = college_code + group_ida
    group.save()
    return encoded

@staff_member_required
def checkout(request,c_id):
    college=get_object_or_404(College,id=c_id)
    if request.method=='POST':
        data=request.POST
        try:
            part_list=Participant.objects.filter(id__in=data.getlist('part_list'))
        except:
            return redirect(request.META.get('HTTP_REFERER'))
        checkout_group=CheckoutGroup.objects.create()
        checkout_group.amount_retained=int(data['retained'])
        checkout_group.save()

        for participant in part_list:
            room=participant.room
            room.vacancy+=1
            room.save()
            participant.checkout_group=checkout_group
            participant.acco=False
            participant.save()

        encoded=generate_ckgroup_code(checkout_group)
        checkout_group.save()
        return redirect(reverse('regsoft:checkout_groups',kwargs={'c_id':college.id}))
    participant_list=Participant.objects.filter(acco=True, college=college)
    return render(request, 'regsoft/checkout.html', {'college':college, 'part_list':participant_list})

@staff_member_required
def master_checkout(request):
    ck_group_list=CheckoutGroup.objects.all()
    rows = [{'data':[ck_group.participant_set.all()[0].college.name,ck_group.participant_set.all().count(), ck_group.created_time, ck_group.amount_retained], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:ck_group_details', kwargs={'ck_id':ck_group.id})), 'title':'View Details'}]} for ck_group in ck_group_list]
    headings = ['College', 'Participant Count', 'Time of Checkout', 'Amount Retained', 'View Details']
    title = 'All Checkout groups'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title,
    }
    amount=0
    for ck_group in CheckoutGroup.objects.all():
        amount+=ck_group.amount_retained
    return render(request, 'regsoft/master_checkout.html', {'tables':[table,], 'amount':amount})

@staff_member_required
def checkout_groups(request, c_id):
    college = get_object_or_404(College, id=c_id)
    ck_group_list = [ck_group for ck_group in CheckoutGroup.objects.all() if ck_group.participant_set.all()[0].college == college]
    rows = [{'data':[ck_group.participant_set.all().count(), ck_group.created_time, ck_group.amount_retained], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:ck_group_details', kwargs={'ck_id':ck_group.id})), 'title':'View Details'}]} for ck_group in ck_group_list]
    headings = ['Participant Count', 'Time of Checkout', 'Amount Retained', 'View Details']
    title = 'Checkout groups from ' + college.name
    table = {
        'rows':rows,
        'headings':headings,
        'title':title,
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

##Helper function to get event strings##
def get_event_string(participant):
    participation_set = MainParticipation.objects.filter(participant=participant, pcr_approved=True)
    events = ''
    for participation in participation_set:
        events += participation.event.name + ', '
    events = events[:-2]
    return events

@staff_member_required
def ck_group_details(request, ck_id):
    checkout_group = get_object_or_404(CheckoutGroup, id=ck_id)
    rows = [{'data':[part.name, part.phone, part.email, part.gender, get_event_string(part), part.room.room, part.room.bhavan.name], 'link':[]} for part in checkout_group.participant_set.all()]
    headings = ['Name', 'Phone', 'Email', 'Gender', 'Events', 'Room', 'Bhavan']
    title = 'Checkout detail at ' + str(checkout_group.created_time) + ', Amount Retained:' + str(checkout_group.amount_retained)
    table = {
        'rows':rows,
        'headings':headings,
        'title':title,
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,],})


################### CONTROLZ ##################

@staff_member_required
def controlz_home(request):
    rows=[]
    for group in Group.objects.all():
        code = group.group_code
        group_leader = get_group_leader(group)
        print(group_leader)
        leader_name = group_leader.name
        leader_college = group_leader.college.name
        leader_phone = group_leader.phone
        time = group.created_time
        no_of_members = group.participant_set.filter(is_guest = False).count()
        controlz_passed = group.participant_set.filter(controlz = True).count()
        bill_url = request.build_absolute_uri(reverse('regsoft:create_bill', kwargs={'g_id':group.id}))
        rows.append({
            'data': [
                code,
                leader_name,
                leader_college,
                leader_phone,
                time,
                no_of_members, 
                controlz_passed
            ],
            'link': [{
                'url':bill_url,
                'title':'Create Bill'
            }]
        })
    headings = ['Group Code', 'Group Leader', 'College', 'Gleader phone', 'Firewallz passed time', 'Total in group', 'Passed controls from group','View Participants']
    title = 'Groups that have passed firewallz'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})


@staff_member_required
def create_bill(request,g_id):
    group=get_object_or_404(Group,id=g_id)
    controlz_passed=group.participant_set.filter(controlz=True,is_guest=False)
    controlz_unpassed=group.participant_set.filter(controlz=False,is_guest=False)
    if request.method=='POST':
        data=request.POST
        try:
            id_list=data.getlist('data')
        except:
            messages.warning(request,'Please select participants')
            return redirect(request.META.get('HTTP_REFERER'))
        if not id_list:
            messages.warning(request, 'Please select participants')
            return redirect(request.META.get('HTTP_REFERER'))
        bill = Bill()
        bill.two_thousands = data['twothousands']
        bill.five_hundreds = data['fivehundreds']
        bill.two_hundreds = data['twohundreds']
        bill.hundreds = data['hundreds']
        bill.fifties = data['fifties']
        bill.twenties = data['twenties']
        bill.tens = data['tens']
        bill.two_thousands_returned = data['twothousandsreturned']
        bill.five_hundreds_returned = data['fivehundredsreturned']
        bill.two_hundreds_returned = data['twohundredsreturned']
        bill.hundreds_returned = data['hundredsreturned']
        bill.fifties_returned = data['fiftiesreturned']
        bill.twenties_returned = data['twentiesreturned']
        bill.tens_returned = data['tensreturned']
        amount_dict = {'twothousands':2000, 'fivehundreds':500, 'twohundreds':200,'hundreds':100, 'fifties':50, 'twenties':20, 'tens':10}
        return_dict = {'twothousandsreturned':2000, 'fivehundredsreturned':500, 'twohundredsreturned':200,'hundredsreturned':100, 'fiftiesreturned':50, 'twentiesreturned':20, 'tensreturned':10}
        bill.amount=0
        for key,value in amount_dict.items():
            bill.amount+=int(data[key])*int(value)
        for key,value in return_dict.items():
            bill.amount-=int(data[key])*int(value)
        try:
            bill.draft_number=data['draft_number']
        except:
            pass
        bill.draft_amount=data['draft_amount']
        bill.amount+=int(bill.draft_amount)
        if not(bill.amount==0 and bill.draft_amount==0):
            bill.save()
            for p_id in id_list:
                part=Participant.objects.get(id=p_id)
                part.bill=bill
                part.controlz=True
                part.curr_controlz_paid=True
                part.curr_paid=True
                part.save()
            return redirect(reverse('regsoft:bill_details', kwargs={'b_id':bill.id}))
        else:
            messages.warning(request,'Please enter a bill amount')
            return redirect(reverse('regsoft:create_bill', kwargs={'g_id':group.id}))
    else:
        return render(request,'regsoft/controlz_group.html',{'controlz_passed':controlz_passed,'controlz_unpassed':controlz_unpassed,'group':group})

@staff_member_required
def show_all_bills(request):
    rows=[{'data':[college.name,college.participant_set.filter(controlz=True).count()],'link':[{'url':request.build_absolute_uri(reverse('regsoft:show_college_bills',kwargs={'c_id':college.id})),'title':'Show bills'}]} for college in College.objects.all()]
    headings=['College','Controlz passed participants','Show bills']
    title='Colleges for bill details'
    table={
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request,'regsoft/tables.html',{'tables':[table,]})

@staff_member_required
def show_college_bills(request,c_id):
    college=get_object_or_404(College,id=c_id)
    bills=[]
    for part in college.participant_set.filter(controlz=True):
        if part.bill not in bills:
            bills.append(part.bill)
    rows = [{'data':[bill.time_paid, bill.participant_set.filter(gender='male').count(), bill.participant_set.filter(gender='female').count(), bill.amount], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:bill_details', kwargs={'b_id':bill.id})), 'title':'View Details'}]} for bill in bills]
    headings = ['Time Created', 'Male Participants', 'Female Participants', 'Amount', 'Details']
    title = 'Bills created for ' + college.name
    table = {
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def bill_details(request,b_id):
    bill=get_object_or_404(Bill,id=b_id)
    from datetime import datetime
    time_stamp=datetime.now()
    participants=bill.participant_set.all()
    college=participants[0].college
    c_rows = [{'data':[part.name, get_event_string(part), bill.time_paid, get_amount(part)], 'link':[]} for part in bill.participant_set.all()]
    table = {
        'title' : 'Participant details for the bill from ' + college.name +'. Cash amount = Rs ' + str(bill.amount-bill.draft_amount) + '. Draft Amount = Rs ' + str(bill.draft_amount),
        'headings' : ['Name', 'Event(s)', 'Time created','Had to pay'],
        'rows':c_rows,
    }
    return render(request, 'regsoft/bill_details.html', {'tables':[table,],'bill':bill, 'participant_list':participants, 'college':college, 'curr_time':time_stamp})

def get_amount(participant):
    if participant.controlz_paid and participant.paid:
        return 0
    elif participant.paid:
        return 750
    else:
        return 1050

@staff_member_required
def print_bill(request,b_id):
    bill=get_object_or_404(Bill,id=b_id)
    from datetime import datetime
    time=datetime.now()
    participants=bill.participant_set.all()
    if not participants:
        return redirect(reverse('regsoft:bill_details', kwargs={'b_id':bill.id}))
    college=participants[0].college
    g_leader=bill.participant_set.all()[0].group.participant_set.get(is_g_leader=True)
    draft=bill.draft_number
    if not draft:
        draft='null'
    payment_methods=[{'method':'Cash','amount':bill.amount-bill.draft_amount},{'method':'Draft number '+draft,'amount':bill.draft_amount}]
    return render(request, 'regsoft/bill_invoice.html', {'bill':bill, 'part_list':participants, 'college':college, 'payment_methods':payment_methods, 'time':time, 'cr':g_leader})

@staff_member_required
def delete_bill(request,b_id):
    bill=get_object_or_404(Bill,id=b_id)
    participants=bill.participant_set.all()
    for part in participants:
        part.controlz=False
        part.curr_controlz_paid=False
        part.curr_paid=False
        part.save()
    bill.delete()
    college=participants[0].college
    return redirect(reverse('regsoft:show_college_bills',kwargs={'c_id':college.id}))
    
@staff_member_required
def recnacc_list(request):
    rows = []
    for group in Group.objects.all().order_by('-created_time'):
        code = group.group_code
        group_leader = get_group_leader(group)
        leader_name = group_leader.name
        leader_college = group_leader.college.name
        leader_phone = group_leader.phone
        time = group.created_time
        controlz_passed = group.participant_set.filter(controlz = True).count()
        total_alloted = group.participant_set.filter(controlz=True, acco=True, checkout_group=None).count()
        checkout_count =  group.participant_set.filter(checkout_group__isnull=False).count()
        acco_details_url = request.build_absolute_uri(reverse('regsoft:recnacc_list_group', kwargs={'g_id':group.id}))
        rows.append({
            'data':[
                code,
                leader_name,
                leader_college,
                leader_phone,
                time,
                controlz_passed,
                total_alloted,
                checkout_count,
                ],
            'link':[{
                'url':acco_details_url,
                'title':'Select Participants'
                }]
            })

    headings = ['Group Code', 'Group Leader', 'College', 'Gleader phone', 'Firewallz passed time', 'Total controlz passed','Total alloted', 'Checkout','View Participants']
    title = 'Groups that have been alloted'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def recnacc_list_group(request, g_id):
    try:
        group = get_object_or_404(Group, id=g_id)
    except:
        response = JsonResponse({'message':'No such group exists'})
        return response
    participant_list = group.participant_set.filter(acco=True).order_by('-recnacc_time')
    context = {'participant_list':participant_list,'college':get_group_leader(group).college}
    return render(request, 'regsoft/recnacc_list.html', context)

@staff_member_required
def generate_recnacc_list(request):
    if request.method == 'POST':
        data = request.POST
        id_list = data.getlist('data')
        c_rows = []
        if len(id_list) == 0:
            context = {
                'error_heading': "Error",
                'message': "Select at least one Participant.",
                'url':request.build_absolute_uri(reverse('regsoft:controlz_home'))
                }
            return render(request, 'registrations/message.html', context)
        for p_id in id_list:
            part = Participant.objects.get(id=p_id)
            c_rows.append({
                'data':[
                    part.name,
                    part.college.name,
                    part.gender,
                    get_cr_name(part), 
                    get_event_string(part), 
                    part.room.room, 
                    part.room.bhavan, 
                    400], 
                'link':[]})
        amount = (len(id_list))*400
        c_rows.append({'data':['Total', '','','','','','',amount]})
        
        table = {
            'title':'Participant list for RecNAcc from ' + part.college.name,
            'headings':['Name', 'College', 'Gender', 'CR Name', 'Event(s)', 'Room','Bhavan', 'Caution Deposit'],
            'rows': c_rows
        }
        return render(request, 'regsoft/tables.html', {'tables':[table,]})

@staff_member_required
def get_profile_card(request):
    rows = [{'data':[part.name, part.phone, part.email, part.gender, get_event_string(part)], 'link':[{'url':request.build_absolute_uri(reverse('regsoft:get_profile_card_participant', kwargs={'p_id':part.id})), 'title':'Get profile card'}]} for part in Participant.objects.filter(Q(pcr_final=True) | Q(is_guest=True))]
    headings = ['Name', 'Phone', 'Email', 'Gender', 'Events', 'Get profile card']
    title = 'Generate Profile Card'
    table = {
        'rows':rows,
        'headings':headings,
        'title':title,
    }
    return render(request, 'regsoft/tables.html', {'tables':[table,],})

@staff_member_required
def get_profile_card_group(request, g_id):
    group = get_object_or_404(Group, id=g_id)
    part_list = group.participant_set.all()
    url = request.build_absolute_uri(reverse('registrations:generate_qr'))
    return render(request, 'regsoft/card.html', {'part_list':part_list, 'url':url})

@staff_member_required
def contacts(request):
    return render(request, 'regsoft/contact.html')

@staff_member_required
def user_logout(request):
    logout(request)
    return redirect('regsoft:index')


# ********************************** INVENTORY VIEWS **********************************

@staff_member_required
def dashboard(request):
    return render(request, 'regsoft/index.html')


@staff_member_required
def dc_login(request):
    if request.method == 'GET':
        dc_list = DC.objects.all()
        return render(request, 'regsoft/dc_login.html',{'dc_list':dc_list})

    if request.method == 'POST':
        print(request.POST)
        try:
            id = request.POST['dc']
        except:
            messages.warning(request,'Please select a club/department')
            return redirect(request.META.get('HTTP_REFERER')) 
        try:
            dc = DC.objects.get(id = id)
        except:
            messages.warning(request,'Invalid unique id')
            return redirect(request.META.get('HTTP_REFERER')) 
        
        return render(request, 'regsoft/dc_home.html', {'dc':dc})

@staff_member_required
def dc_new_entry(request, dc_id):
    if request.method == 'GET':
        try:
            dc = DC.objects.get(id = dc_id)
        except:
            context = {
                'error_heading': "Error",
                'message': "Club/Department does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
                }
            return render(request, 'registrations/message.html', context)
        return render(request, 'regsoft/dc_new_form.html', context={'dc':dc})

    else:
        data = request.POST
        print(data)
        try:
            dc = DC.objects.get(id = dc_id)
        except:
            context = {
                'error_heading': "Error",
                'message': "Club/Department does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
                }
            return render(request, 'registrations/message.html', context)
        try:
            inventory = DC_Inventory.objects.get(dc=dc)
            inventory.comments = data['comment']
            try:
                inventory.blankets += int(data['blankets'])
                inventory.mattress += int(data['mattresses'])
                inventory.pillows += int(data['pillows'])
                inventory.spikes += int(data['spikes'])
                inventory.bedsheets += int(data['bedsheets'])
                inventory.quilts += int(data['quilts'])
                inventory.buckets += int(data['buckets'])
                inventory.mugs += int(data['mugs'])
                inventory.fans += int(data['fans'])
                inventory.bulbs += int(data['bulbs'])
                inventory.water_campers += int(data['watercampers'])
                inventory.water_drums += int(data['waterdrums'])
                inventory.waste_drums += int(data['waste_drums'])
                inventory.tables += int(data['tables'])
                inventory.table_cloths += int(data['table_cloths'])
                inventory.chairs += int(data['chairs'])
                inventory.red_carpets += int(data['red_carpets'])
                inventory.green_carpets += int(data['green_carpets'])
                inventory.curtains += int(data['curtains'])
                inventory.halogen_lamps += int(data['halogen_lamps'])
                inventory.sodium_lamps += int(data['sodium_lamps'])
                inventory.iron_poles += int(data['iron_poles'])
                inventory.paper_rolls += int(data['paper_rolls'])
                inventory.bamboo_poles += int(data['bamboo_poles'])
                inventory.ropes += int(data['ropes'])
                inventory.wires += int(data['wires'])
                inventory.tents += int(data['tents'])
                inventory.item1 += int(data['item1'])
                inventory.item2 += int(data['item2'])
                inventory.item3 += int(data['item3'])
                inventory.item4 += int(data['item4'])
                inventory.item5 += int(data['item5'])
                inventory.save()
            except Exception as e:
                print(e)
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER')) 

        except:
            inventory = DC_Inventory.objects.create(dc = dc)
            inventory.comments = data['comment']
            try:
                inventory.blankets = int(data['blankets'])
                inventory.mattress = int(data['mattresses'])
                inventory.pillows = int(data['pillows'])
                inventory.spikes = int(data['spikes'])
                inventory.bedsheets = int(data['bedsheets'])
                inventory.quilts = int(data['quilts'])
                inventory.buckets = int(data['buckets'])
                inventory.mugs = int(data['mugs'])
                inventory.fans = int(data['fans'])
                inventory.bulbs = int(data['bulbs'])
                inventory.water_campers = int(data['watercampers'])
                inventory.water_drums = int(data['waterdrums'])
                inventory.waste_drums = int(data['waste_drums'])
                inventory.tables = int(data['tables'])
                inventory.table_cloths = int(data['table_cloths'])
                inventory.chairs = int(data['chairs'])
                inventory.red_carpets = int(data['red_carpets'])
                inventory.green_carpets = int(data['green_carpets'])
                inventory.curtains = int(data['curtains'])
                inventory.halogen_lamps = int(data['halogen_lamps'])
                inventory.sodium_lamps = int(data['sodium_lamps'])
                inventory.iron_poles = int(data['iron_poles'])
                inventory.paper_rolls = int(data['paper_rolls'])
                inventory.bamboo_poles = int(data['bamboo_poles'])
                inventory.ropes = int(data['ropes'])
                inventory.wires = int(data['wires'])
                inventory.tents = int(data['tents'])
                inventory.item1 = int(data['item1'])
                inventory.item2 = int(data['item2'])
                inventory.item3 = int(data['item3'])
                inventory.item4 = int(data['item4'])
                inventory.item5 = int(data['item5'])
                inventory.save()
            except Exception as e:
                print(e)
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER')) 


        messages.warning(request,'Successfully Added')
        return redirect(request.META.get('HTTP_REFERER')) 

@staff_member_required
def dc_remove_entry(request, dc_id):
    if request.method == 'GET':
        try:
            dc = DC.objects.get(id = dc_id)
        except:
            context = {
                'error_heading': "Error",
                'message': "Club/Department does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
                }
            return render(request, 'registrations/message.html', context)
        try:
            inventory = DC_Inventory.objects.get(dc = dc)
        except:
            messages.warning(request,'Inventory for the selected option does not exist')
            return redirect(request.META.get('HTTP_REFERER'))
        return render(request, 'regsoft/dc_remove_form.html', context={'dc':dc,'inventory':inventory})

    else:
        data = request.POST
        try:
            dc = DC.objects.get(id = dc_id)
        except:
            context = {
                'error_heading': "Error",
                'message': "Club/Department does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
                }
            return render(request, 'registrations/message.html', context)
        try:
            inventory = DC_Inventory.objects.get(dc=dc)
            inventory.comments = data['comment']
            try:
                inventory.blankets -= int(data['blankets'])
                inventory.mattress -= int(data['mattresses'])
                inventory.pillows -= int(data['pillows'])
                inventory.spikes -= int(data['spikes'])
                inventory.bedsheets -= int(data['bedsheets'])
                inventory.quilts -= int(data['quilts'])
                inventory.buckets -= int(data['buckets'])
                inventory.mugs -= int(data['mugs'])
                inventory.fans -= int(data['fans'])
                inventory.bulbs -= int(data['bulbs'])
                inventory.water_campers -= int(data['watercampers'])
                inventory.water_drums -= int(data['waterdrums'])
                inventory.waste_drums -= int(data['waste_drums'])
                inventory.tables -= int(data['tables'])
                inventory.table_cloths -= int(data['table_cloths'])
                inventory.chairs -= int(data['chairs'])
                inventory.red_carpets -= int(data['red_carpets'])
                inventory.green_carpets -= int(data['green_carpets'])
                inventory.curtains -= int(data['curtains'])
                inventory.halogen_lamps -= int(data['halogen_lamps'])
                inventory.sodium_lamps -= int(data['sodium_lamps'])
                inventory.iron_poles -= int(data['iron_poles'])
                inventory.paper_rolls -= int(data['paper_rolls'])
                inventory.bamboo_poles -= int(data['bamboo_poles'])
                inventory.ropes -= int(data['ropes'])
                inventory.wires -= int(data['wires'])
                inventory.tents -= int(data['tents'])
                inventory.item1 -= int(data['item1'])
                inventory.item2 -= int(data['item2'])
                inventory.item3 -= int(data['item3'])
                inventory.item4 -= int(data['item4'])
                inventory.item5 -= int(data['item5'])
                inventory.save()
            
            except:
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER'))
        
        except:
            context = {
                'error_heading': "Error",
                'message': "Inventory does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
                }
            return render(request, 'registrations/message.html', context)
        
        messages.warning(request,'Successfully Returned')
        return redirect(request.META.get('HTTP_REFERER'))


@staff_member_required
def dc_view_status(request, dc_id):
    try:
        dc = DC.objects.get(id = dc_id)
    except:
        context = {
            'error_heading': "Error",
            'message': "Club/Department does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
            }
        return render(request, 'registrations/message.html', context)
    try:
        inventory = DC_Inventory.objects.get(dc = dc)
    except:
        context = {
            'error_heading': "No Inventory",
            'message': "Inventory does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:dc_login'))
            }
        return render(request, 'registrations/message.html', context)

    context = {'inventory':inventory}
    return render(request, 'regsoft/view_dc.html',context)


@staff_member_required
def tender_home(request):
    return render(request, 'regsoft/tender_home.html',{'name':'Tender'})

@staff_member_required
def tender_new_entry(request):
    if request.method == 'GET':
        try:
            loc = Location.objects.filter(tender = True)
        except Exception as e:
            print(e)
            messages.warning(request,'Invalid locations')
            return redirect(request.META.get('HTTP_REFERER'))
        return render(request, 'regsoft/tender_new_form.html', {'locations':loc,'name':"Tender"})

    else:
        data = request.POST
        print(data)
        try:
            loc = Location.objects.get(id = data['location'])
        except Exception as e:
            print(e)
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location = loc)
            inventory.comments = data['comment']
            try:
                inventory.blankets += int(data['blankets'])
                inventory.mattress += int(data['mattresses'])
                inventory.pillows += int(data['pillows'])
                inventory.spikes += int(data['spikes'])
                inventory.bedsheets += int(data['bedsheets'])
                inventory.quilts += int(data['quilts'])
                inventory.buckets += int(data['buckets'])
                inventory.mugs += int(data['mugs'])
                inventory.fans += int(data['fans'])
                inventory.bulbs += int(data['bulbs'])
                inventory.water_campers += int(data['watercampers'])
                inventory.water_drums += int(data['waterdrums'])
                inventory.waste_drums += int(data['waste_drums'])
                inventory.tables += int(data['tables'])
                inventory.table_cloths += int(data['table_cloths'])
                inventory.chairs += int(data['chairs'])
                inventory.red_carpets += int(data['red_carpets'])
                inventory.green_carpets += int(data['green_carpets'])
                inventory.curtains += int(data['curtains'])
                inventory.halogen_lamps += int(data['halogen_lamps'])
                inventory.sodium_lamps += int(data['sodium_lamps'])
                inventory.iron_poles += int(data['iron_poles'])
                inventory.paper_rolls += int(data['paper_rolls'])
                inventory.bamboo_poles += int(data['bamboo_poles'])
                inventory.ropes += int(data['ropes'])
                inventory.wires += int(data['wires'])
                inventory.tents += int(data['tents'])
                inventory.item1 += int(data['item1'])
                inventory.item2 += int(data['item2'])
                inventory.item3 += int(data['item3'])
                inventory.item4 += int(data['item4'])
                inventory.item5 += int(data['item5'])
                inventory.save()
            except Exception as e:
                print(e)
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER'))
        except:
            inventory = Inventory.objects.create(location = loc)
            inventory.comments = data['comment']
            try:
                inventory.blankets = int(data['blankets'])
                inventory.mattress = int(data['mattresses'])
                inventory.pillows = int(data['pillows'])
                inventory.spikes = int(data['spikes'])
                inventory.bedsheets = int(data['bedsheets'])
                inventory.quilts = int(data['quilts'])
                inventory.buckets = int(data['buckets'])
                inventory.mugs = int(data['mugs'])
                inventory.fans = int(data['fans'])
                inventory.bulbs = int(data['bulbs'])
                inventory.water_campers = int(data['watercampers'])
                inventory.water_drums = int(data['waterdrums'])
                inventory.waste_drums = int(data['waste_drums'])
                inventory.tables = int(data['tables'])
                inventory.table_cloths = int(data['table_cloths'])
                inventory.chairs = int(data['chairs'])
                inventory.red_carpets = int(data['red_carpets'])
                inventory.green_carpets = int(data['green_carpets'])
                inventory.curtains = int(data['curtains'])
                inventory.halogen_lamps = int(data['halogen_lamps'])
                inventory.sodium_lamps = int(data['sodium_lamps'])
                inventory.iron_poles = int(data['iron_poles'])
                inventory.paper_rolls = int(data['paper_rolls'])
                inventory.bamboo_poles = int(data['bamboo_poles'])
                inventory.ropes = int(data['ropes'])
                inventory.wires = int(data['wires'])
                inventory.tents = int(data['tents'])
                inventory.item1 = int(data['item1'])
                inventory.item2 = int(data['item2'])
                inventory.item3 = int(data['item3'])
                inventory.item4 = int(data['item4'])
                inventory.item5 = int(data['item5'])
                inventory.save()
            except:
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER')) 

        messages.warning(request,'Successfully Added')
        return redirect(request.META.get('HTTP_REFERER'))

@staff_member_required
def tender_remove_home(request):
    loc = Location.objects.filter(tender = True)
    return render(request, 'regsoft/return_home.html',{'loc_list':loc})

def tender_in(request):
    if request.method == 'POST':
        try:
            id = request.POST['location']
        except Exception as e:
            print(e)
            context = {
            'error_heading': "Error",
            'message': "Location does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:tender_home'))
            }
            return render(request, 'registrations/message.html', context)
        try:
            inventory = Inventory.objects.get(location = Location.objects.get(id=id))
        except Exception as e:
            print(e)
            context = {
                'error_heading': "Error",
                'message': "Inventory does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:tender_home'))
            }
            return render(request, 'registrations/message.html', context)
        return redirect('regsoft:tender_remove_entry', l_id=id )


@staff_member_required
def tender_remove_entry(request, l_id):
    if request.method == 'GET':
        try:
            loc = Location.objects.filter(id = l_id)
        except:
            context = {
            'error_heading': "Error",
            'message': "Location does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:tender_home'))
            }
            return render(request, 'registrations/message.html', context)
        try:
            inventory = Inventory.objects.get(location = loc)
        except:
            context = {
            'error_heading': "Error",
            'message': "Inventory does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:tender_home'))
            }
            return render(request, 'registrations/message.html', context)
        return render(request, 'regsoft/tender_remove_form.html', {'locations':loc,'inventory':inventory,'id':l_id})
    else:
        data = request.POST
        try:
            loc = Location.objects.get(id = l_id)
        except:
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location=loc)
            try:
                inventory.blankets -= int(data['blankets'])
                inventory.mattress -= int(data['mattresses'])
                inventory.pillows -= int(data['pillows'])
                inventory.spikes -= int(data['spikes'])
                inventory.bedsheets -= int(data['bedsheets'])
                inventory.quilts -= int(data['quilts'])
                inventory.buckets -= int(data['buckets'])
                inventory.mugs -= int(data['mugs'])
                inventory.fans -= int(data['fans'])
                inventory.bulbs -= int(data['bulbs'])
                inventory.water_campers -= int(data['watercampers'])
                inventory.water_drums -= int(data['waterdrums'])
                inventory.waste_drums -= int(data['waste_drums'])
                inventory.tables -= int(data['tables'])
                inventory.table_cloths -= int(data['table_cloths'])
                inventory.chairs -= int(data['chairs'])
                inventory.red_carpets -= int(data['red_carpets'])
                inventory.green_carpets -= int(data['green_carpets'])
                inventory.curtains -= int(data['curtains'])
                inventory.halogen_lamps -= int(data['halogen_lamps'])
                inventory.sodium_lamps -= int(data['sodium_lamps'])
                inventory.iron_poles -= int(data['iron_poles'])
                inventory.paper_rolls -= int(data['paper_rolls'])
                inventory.bamboo_poles -= int(data['bamboo_poles'])
                inventory.ropes -= int(data['ropes'])
                inventory.wires -= int(data['wires'])
                inventory.tents -= int(data['tents'])
                inventory.item1 -= int(data['item1'])
                inventory.item2 -= int(data['item2'])
                inventory.item3 -= int(data['item3'])
                inventory.item4 -= int(data['item4'])
                inventory.item5 -= int(data['item5'])
                inventory.save()
            except:
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER'))
        
        except Exception as e:
            print(e)
            context = {
                'error_heading': "Error",
                'message': "Inventory does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:tender_home'))
                }
            return render(request, 'registrations/message.html', context)

        messages.warning(request,'Successfully Updated')
        return redirect(request.META.get('HTTP_REFERER'))


@staff_member_required
def tender_view_status(request):
    if request.method == 'GET':
        loc = Location.objects.filter(tender = True)
        return render(request, 'regsoft/tender_view.html', {'locations':loc,'name':'tender'})
    else:
        try:
            loc = Location.objects.get(id = request.POST['location'])
        except:
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location = loc)
            return render(request, 'regsoft/tender_view_status.html', {'inventory':inventory})
        except:
            messages.warning(request,'Inventory does not exist')
            return redirect(request.META.get('HTTP_REFERER'))

@staff_member_required
def mattress_home(request):
    return render(request, 'regsoft/mattress_home.html', {'name':'Mattress'})

@staff_member_required
def mattress_new_entry(request):
    if request.method == 'GET':
        loc = Location.objects.filter(tender = False)
        return render(request, 'regsoft/mattress_new_form.html', {'locations':loc,'name':'Mattress'})

    else:
        data = request.POST
        try:
            loc = Location.objects.get(id = request.POST['location'])
        except Exception as e:
            print(e)
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location = loc)
            inventory.comments = data['comment']
            try:
                inventory.blankets += int(data['blankets'])
                inventory.mattress += int(data['mattresses'])
                inventory.pillows += int(data['pillows'])
                inventory.spikes += int(data['spikes'])
                inventory.bedsheets += int(data['bedsheets'])
                inventory.quilts += int(data['quilts'])
                inventory.buckets += int(data['buckets'])
                inventory.mugs += int(data['mugs'])
                inventory.fans += int(data['fans'])
                inventory.bulbs += int(data['bulbs'])
                inventory.water_campers += int(data['watercampers'])
                inventory.water_drums += int(data['waterdrums'])
                inventory.waste_drums += int(data['waste_drums'])
                inventory.tables += int(data['tables'])
                inventory.table_cloths += int(data['table_cloths'])
                inventory.chairs += int(data['chairs'])
                inventory.red_carpets += int(data['red_carpets'])
                inventory.green_carpets += int(data['green_carpets'])
                inventory.curtains += int(data['curtains'])
                inventory.halogen_lamps += int(data['halogen_lamps'])
                inventory.sodium_lamps += int(data['sodium_lamps'])
                inventory.iron_poles += int(data['iron_poles'])
                inventory.paper_rolls += int(data['paper_rolls'])
                inventory.bamboo_poles += int(data['bamboo_poles'])
                inventory.ropes += int(data['ropes'])
                inventory.wires += int(data['wires'])
                inventory.tents += int(data['tents'])
                inventory.item1 += int(data['item1'])
                inventory.item2 += int(data['item2'])
                inventory.item3 += int(data['item3'])
                inventory.item4 += int(data['item4'])
                inventory.item5 += int(data['item5'])
                inventory.save()
            except:
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER'))
        except:
            inventory = Inventory.objects.create(location = loc)
            inventory.comments = data['comment']
            try:
                inventory.blankets = int(data['blankets'])
                inventory.mattress = int(data['mattresses'])
                inventory.pillows = int(data['pillows'])
                inventory.spikes = int(data['spikes'])
                inventory.bedsheets = int(data['bedsheets'])
                inventory.quilts = int(data['quilts'])
                inventory.buckets = int(data['buckets'])
                inventory.mugs = int(data['mugs'])
                inventory.fans = int(data['fans'])
                inventory.bulbs = int(data['bulbs'])
                inventory.water_campers = int(data['watercampers'])
                inventory.water_drums = int(data['water_drums'])
                inventory.waste_drums = int(data['waste_drums'])
                inventory.tables = int(data['tables'])
                inventory.table_cloths = int(data['table_cloths'])
                inventory.chairs = int(data['chairs'])
                inventory.red_carpets = int(data['red_carpets'])
                inventory.green_carpets = int(data['green_carpets'])
                inventory.curtains = int(data['curtains'])
                inventory.halogen_lamps = int(data['halogen_lamps'])
                inventory.sodium_lamps = int(data['sodium_lamps'])
                inventory.iron_poles = int(data['iron_poles'])
                inventory.paper_rolls = int(data['paper_rolls'])
                inventory.bamboo_poles = int(data['bamboo_poles'])
                inventory.ropes = int(data['ropes'])
                inventory.wires = int(data['wires'])
                inventory.tents = int(data['tents'])
                inventory.item1 = int(data['item1'])
                inventory.item2 = int(data['item2'])
                inventory.item3 = int(data['item3'])
                inventory.item4 = int(data['item4'])
                inventory.item5 = int(data['item5'])
                inventory.save()
            except:
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER')) 

        messages.warning(request,'Successfully Added')
        return redirect(request.META.get('HTTP_REFERER'))

@staff_member_required
def mattress_return_home(request):
    loc = Location.objects.filter(tender = False)
    return render(request, 'regsoft/return_mattress.html',{'loc_list':loc})

@staff_member_required
def mattress_in(request):
    try:
        id = request.POST['location']
    except:
        context = {
            'error_heading': "Error",
            'message': "Location does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:mattress_home'))
            }
        return render(request, 'registrations/message.html', context)
    return redirect('regsoft:mattress_remove_entry', l_id=id)


@staff_member_required
def mattress_remove_entry(request,l_id):
    if request.method == 'GET':
        loc = Location.objects.filter(tender = False)
        try:
            inventory = Inventory.objects.get(id=l_id)
        except:
            context = {
            'error_heading': "Error",
            'message': "Location does not exist",
            'url':request.build_absolute_uri(reverse('regsoft:mattress_home'))
            }
            return render(request, 'registrations/message.html', context)
        return render(request, 'regsoft/mattress_remove_form.html', {'locations':loc,'inventory':inventory,'id':l_id})
    else:
        data = request.POST
        try:
            loc = Location.objects.get(id = data['location'])
        except:
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location = loc)
            inventory.comments = data['comment']
            try:
                inventory.blankets -= int(data['blankets'])
                inventory.mattress -= int(data['mattresses'])
                inventory.pillows -= int(data['pillows'])
                inventory.spikes -= int(data['spikes'])
                inventory.bedsheets -= int(data['bedsheets'])
                inventory.quilts -= int(data['quilts'])
                inventory.buckets -= int(data['buckets'])
                inventory.mugs -= int(data['mugs'])
                inventory.fans -= int(data['fans'])
                inventory.bulbs -= int(data['bulbs'])
                inventory.water_campers -= int(data['watercampers'])
                inventory.water_drums -= int(data['waterdrums'])
                inventory.waste_drums -= int(data['waste_drums'])
                inventory.tables -= int(data['tables'])
                inventory.table_cloths -= int(data['table_cloths'])
                inventory.chairs -= int(data['chairs'])
                inventory.red_carpets -= int(data['red_carpets'])
                inventory.green_carpets -= int(data['green_carpets'])
                inventory.curtains -= int(data['curtains'])
                inventory.halogen_lamps -= int(data['halogen_lamps'])
                inventory.sodium_lamps -= int(data['sodium_lamps'])
                inventory.iron_poles -= int(data['iron_poles'])
                inventory.paper_rolls -= int(data['paper_rolls'])
                inventory.bamboo_poles -= int(data['bamboo_poles'])
                inventory.ropes -= int(data['ropes'])
                inventory.wires -= int(data['wires'])
                inventory.tents -= int(data['tents'])
                inventory.item1 -= int(data['item1'])
                inventory.item2 -= int(data['item2'])
                inventory.item3 -= int(data['item3'])
                inventory.item4 -= int(data['item4'])
                inventory.item5 -= int(data['item5'])
                inventory.save()
            except Exception as e:
                print(e)
                messages.warning(request,'Please Enter data in proper format')
                return redirect(request.META.get('HTTP_REFERER'))
        
        except:
            context = {
                'error_heading': "Error",
                'message': "Inventory does not exist",
                'url':request.build_absolute_uri(reverse('regsoft:mattress_home'))
                }
            return render(request, 'registrations/message.html', context)
        
        messages.warning(request,'Successfully Updated')
        return redirect(request.META.get('HTTP_REFERER'))

@staff_member_required
def mattress_view_status(request):
    if request.method == 'GET':
        loc = Location.objects.filter(tender = False)
        return render(request, 'regsoft/tender_view.html', {'locations':loc,'name':'Mattress'})
    else:
        try:
            loc = Location.objects.get(id = request.POST['location'])
        except:
            messages.warning(request,'Invalid location')
            return redirect(request.META.get('HTTP_REFERER'))
        try:
            inventory = Inventory.objects.get(location = Location)
            return render(request, 'regsoft/tender_view_status.html', {'location':loc})
            
        except:
            messages.warning(request,'Inventory does not exist')
            return redirect(request.META.get('HTTP_REFERER'))
def tender_new_form(request):
    return HttpResponse('hello')

def excel(request):
    from openpyxl import Workbook
    wb = Workbook(write_only = True)
    ws = wb.create_sheet()
    headings = [
        'Location',
        'Comments',
        'Blankets',
        'Mattress',
        'Pillows',
        'Spikes',
        'Bedsheets',
        'Quilts',
        'Buckets',
        'Mugs',
        'Fans',
        'Bulbs',
        'water_campers',
        'water_drums',
        'waste_drums',
        'Tables',
        'Table Cloths',
        'Chairs',
        'Curtains',
        'Halogen lamps',
        'Sodium Lamps',
        'Tents',
        'Iron Poles',
        'Paper Rolls',
        'Bamboo Poles',
        'Ropes',
        'Wires',
        'Item 1',
        'Item 2',
        'Item 3',
        'Item 4',
        'Item 5'
        ]
    ws.append(headings)
    for location in Location.objects.all():
        try:
            inventory = Inventory.objects.get(location=location)
        except:
            continue
        items = [
            location.name,
            inventory.comments,
            inventory.blankets,
            inventory.mattress,
            inventory.pillows,
            inventory.spikes,
            inventory.bedsheets,
            inventory.quilts,
            inventory.buckets,
            inventory.mugs,
            inventory.fans,
            inventory.bulbs,
            inventory.water_campers,
            inventory.water_drums,
            inventory.waste_drums,
            inventory.tables,
            inventory.table_cloths,
            inventory.chairs,
            inventory.curtains,
            inventory.halogen_lamps,
            inventory.sodium_lamps,
            inventory.tents,
            inventory.iron_poles,
            inventory.paper_rolls,
            inventory.bamboo_poles,
            inventory.ropes,
            inventory.wires,
            inventory.item1,
            inventory.item2,
            inventory.item3,
            inventory.item4,
            inventory.item5
        ]
        ws.append(items)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    wb.save(response)

    return response

def excel2(request):
    from shop.models.item import Tickets
    from openpyxl import Workbook
    wb = Workbook(write_only = True)
    ws = wb.create_sheet()
    headings = [
        'Name',
        'EMail',
        'College',
        'No. of Tickets',
        'Qr code number',
        'link'
        ]
    ws.append(headings)

    prof_show = MainProfShow.objects.get(name__icontains='Guthrie')
    tickets = Tickets.objects.filter(prof_show=prof_show)
    users = [ticket.user for ticket in tickets]
    a=1

    for user,ticket in zip(users,tickets):
        try:
            p=Bitsian.objects.get(user = user)
            bitsian = True
        except:
            p=Participant.objects.get(user=user)
            bitsian=False
        if bitsian:
            college = "BITS"
        else:
            college = p.college.name
        try:
            link = 'https://bits-oasis.org/2018/storewebapp/qr/'+str(user.wallet.uuid)
        except:
            link=''
        li=[p.name,p.email,college,ticket.count,a,link]
        a+=ticket.count
        ws.append(li)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mydata.xlsx'

    wb.save(response)

    return response


def indo_excel(request):
    from shop.models.item import Tickets
    from openpyxl import Workbook
    wb = Workbook(write_only = True)
    ws = wb.create_sheet()
    headings = [
        'Name',
        'Email',
        'College',
        'No. of Tickets',
        'Qr code number',
        'link'
        ]
    ws.append(headings)

    prof_show = MainProfShow.objects.get(name__icontains='Indosoul')
    tickets = Tickets.objects.filter(prof_show=prof_show)
    users = [ticket.user for ticket in tickets]
    a=1

    for user,ticket in zip(users,tickets):
        try:
            p=Bitsian.objects.get(user = user)
            bitsian = True
        except:
            p=Participant.objects.get(user=user)
            bitsian=False
        if bitsian:
            college = "BITS"
        else:
            college = p.college.name
        try:
            link = 'https://bits-oasis.org/2018/storewebapp/qr/'+str(user.wallet.uuid)
        except:
            link=''
        li=[p.name,p.email,college,ticket.count,ticket.qr_no,link]
        a+=ticket.count
        ws.append(li)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Indo_data.xlsx'

    wb.save(response)

    return response



def shankar_excel(request):
    from shop.models.item import Tickets
    from openpyxl import Workbook
    wb = Workbook(write_only = True)
    ws = wb.create_sheet()
    headings = [
        'Name',
        'EMail',
        'College',
        'No. of Tickets',
        'Qr code number',
        'link'
        ]
    ws.append(headings)

    prof_show = MainProfShow.objects.get(name__icontains='Shankar')
    tickets = Tickets.objects.filter(prof_show=prof_show)
    users = [ticket.user for ticket in tickets]
    a=1

    for user,ticket in zip(users,tickets):
        try:
            p=Bitsian.objects.get(user = user)
            bitsian = True
        except:
            p=Participant.objects.get(user=user)
            bitsian=False
        if bitsian:
            college = "BITS"
        else:
            college = p.college.name
        try:
            link = 'https://bits-oasis.org/2018/storewebapp/qr/'+str(user.wallet.uuid)
        except:
            link=''
        li=[p.name,p.email,college,ticket.count,ticket.qr_no,link]
        a+=ticket.count
        ws.append(li)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Shankar_data.xlsx'

    wb.save(response)

    return response