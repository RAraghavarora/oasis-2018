from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from openpyxl import load_workbook
import csv

filename = BASE_DIR + "/utils/student_list3.xlsx"
filename2='/home/sanchit/Downloads/student_list2.xlsx'
wb = load_workbook(filename=filename)
sheet = wb["MASTERSHEET"]

length = len(tuple(sheet.rows))

for i in range(2, length):

	un= sheet.cell(row=i,column=1).value
	long_id = sheet.cell(row=i, column=2).value
	email = sheet.cell(row=i, column=10).value
	name = sheet.cell(row=i, column=4).value.title()
	room_no = int(sheet.cell(row=i, column=5).value)
	bhawan = sheet.cell(row=i, column=6).value
	if email[1:5]=='2017' or email[1:5]=='2018':
		continue
	print(long_id, email, name, room_no, bhawan)

	try:
		user=User.objects.get(username=un[0:9].lower())
		user.wallet.balance.delete()
		print('deleted')
		try:
			user.wallet.delete()
		except:
			pass
		user.delete
		user=User.objects.create(username=email.split('@')[0])	
	except:
		try:
			user=User.objects.create(username=email.split('@')[0])
		except:
			print('user\t',user)
	try:
		Bitsian.objects.get(user=user, name=name, long_id=long_id, room_no=room_no, bhawan=bhawan, email=email)
	except:
		Bitsian.objects.create(user=user, name=name, long_id=long_id, room_no=room_no, bhawan=bhawan, email=email)
