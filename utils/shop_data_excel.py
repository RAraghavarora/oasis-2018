from openpyxl import *
from shop.models import *
from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User

filename = BASE_DIR + "/utils/student_list2.xlsx"
filename2='/home/sanchit/Downloads/STALLS MENU.xlsx'
wb = load_workbook(filename=filename2)
sheet = wb["Sheet1"]
length = len(tuple(sheet.rows))
stall_list=[]
for i in range(2,length):
    stall_name=sheet.cell(row=i,column=1).value
    item_name=sheet.cell(row=i,column=2).value
    price=sheet.cell(row=i,column=3).value
    stall_list.append(stall_name)
    
    # user, created = User.objects.get_or_create(username=row[-1])

    
