from oasis2018.settings import BASE_DIR
from registrations.models import Bitsian
from django.contrib.auth.models import User
from openpyxl import load_workbook
import csv

filename = BASE_DIR + "/utils/student_list2.xlsx"
filename2='/home/sanchit/Downloads/student_list2.xlsx'
wb = load_workbook(filename=filename)
sheet = wb["MASTERSHEET"]

length = len(tuple(sheet.rows))

li1 = list()
li2 = list()
for i in range(2, length+1):

	long_id = sheet.cell(row=i, column=2).value
	other_id = sheet.cell(row=i, column=1).value
	e = sheet.cell(row=i,column=3).value
	if int(long_id[:4]) == 2017 or int(long_id[:4]) == 2018 :
		print(e)
	else:
		email = other_id[0].lower() + long_id[:4] + long_id[9:] + '@pilani.bits-pilani.ac.in'
		print(email)