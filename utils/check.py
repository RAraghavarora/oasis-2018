from openpyxl import load_workbook
filename2='/home/sanchit/Downloads/student_list2.xlsx'
wb = load_workbook(filename=filename2)
sheet = wb["MASTERSHEET"]

length = len(tuple(sheet.rows))
long_id_list=[]
email_list=[]
row=2
for i in range(2,length):
    long_id=sheet.cell(row=i,column=2).value
    email=sheet.cell(row=i,column=10).value
    long_id_list.append(long_id)
    email_list.append(email)

long_id_set=len(set(long_id_list))
email_list_set=len(set(email_list))
print ("long_id"+str(long_id_set))
print("email"+str(email_list_set))


    